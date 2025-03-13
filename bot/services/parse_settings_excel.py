import openpyxl
from aiogram.types.input_file import FSInputFile
from openpyxl.worksheet.worksheet import Worksheet

from bot.models.camera import Camera
from repository import CamerasRepository, CrossingConfigRepository, CrossingConfigButtonsRepository
from bot.models.crossing_config import CrossingMode
from bot.models.template import Template
from bot.models.crossing_config_buttons import CrossingConfigButtons
from repository.templates_repository import TemplatesRepository
from settings import get_settings
from repository.messages import MessagesRepository
from bot.models.messages import Message


class ExcelSettings:
    def __init__(self):
        settings = get_settings()
        self.file_path = settings.SETTINGS_FILE_PATH
        self.cameras_repository = CamerasRepository()
        self.crossing_config_repository = CrossingConfigRepository()
        self.templates_repository = TemplatesRepository()
        self.crossing_config_buttons_repository = CrossingConfigButtonsRepository()

    async def parse_and_save(self) -> tuple[bool, str]:
        try:
            wb_obj = openpyxl.load_workbook(self.file_path)
            for sheet_name in wb_obj.sheetnames:
                sheet_obj = wb_obj[sheet_name]
                if sheet_name.strip() == "Камеры":
                    await self.save_cameras(sheet_obj)
                elif sheet_name.strip() == "Настройки переправы":
                    await self.save_crossing_config(sheet_obj)
                elif sheet_name.strip() == "Шаблоны уведомлений":
                    await self.save_static_message_templates(sheet_obj)
                elif sheet_name.strip() == "Сообщения":
                    await self.save_messages(sheet_obj)
            return True, "OK"
        except Exception as e:
            return False, str(e)

    async def save_crossing_config(self, sheet_obj: Worksheet):
        rows = list(sheet_obj.iter_rows())
        crossing_mode = CrossingMode.get_crossing_mode_by_name(int(rows[1][1].value))
        await self.crossing_config_repository.update_crossing_config(**{"crossing_mode": crossing_mode})
        await self.crossing_config_buttons_repository.delete_all_crossing_config_buttons()

        for row in sheet_obj.iter_rows(min_row=5, values_only=True):
            crossing_config_button = CrossingConfigButtons(
                button_name=row[0],
                button_value=row[1],
                crossing_mode=CrossingMode.get_crossing_mode_by_name(int(row[2]))
            )
            await self.crossing_config_buttons_repository.insert_crossing_config_button(crossing_config_button)

    async def save_messages(self, sheet_obj: Worksheet):
        messages_repository = MessagesRepository()
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            message = Message(key=row[0], name=row[1], text=row[2])
            await messages_repository.create_message(message)

    async def save_cameras(self, sheet_obj: Worksheet):
        current_cameras = await self.cameras_repository.get_all_cameras()
        current_cameras_names = [camera.name for camera in current_cameras]
        config_cameras_names = []
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            camera = Camera(name=row[0], camera_url=row[1])
            config_cameras_names.append(camera.name)
            await self.cameras_repository.create_or_update_camera(camera)

        for current_camera_name in current_cameras_names:
            if current_camera_name not in current_cameras_names:
                await self.cameras_repository.delete_camera(current_camera_name)

    async def save_static_message_templates(self, sheet_obj: Worksheet):
        await self.templates_repository.delete_all_templates()
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            template = Template(
                crossing_mode=CrossingMode.get_crossing_mode_by_name(int(row[0])),
                button_name=row[1], message=row[2], buttons_list=row[3]
            )
            await self.templates_repository.create_template(template)

    async def get_excel_file(self):
        return FSInputFile(self.file_path)
